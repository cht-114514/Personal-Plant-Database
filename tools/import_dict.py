#!/usr/bin/env python3
"""
植物资料库 - 拉丁学名词典导入脚本

使用方法:
  python tools/import_dict.py path/to/dictionary.xlsx
  python tools/import_dict.py --dry-run path/to/dictionary.xlsx
"""

import sys
import sqlite3
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / 'data' / 'botanical.db'


def init_dict_table(conn):
    """创建词典表"""
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS dictionary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            latin_term TEXT NOT NULL,
            chinese_meaning TEXT,
            english_meaning TEXT,
            pronunciation TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_dict_latin ON dictionary(latin_term);
    ''')


def detect_columns(headers):
    """从表头自动检测列映射"""
    col_map = {
        'latin_term': None,
        'chinese_meaning': None,
        'english_meaning': None,
        'pronunciation': None,
    }

    latin_keywords = ['拉丁', 'latin', '词', 'term', '词条', '加词']
    chinese_keywords = ['中文', 'chinese', '释义', '中文释义']
    english_keywords = ['英文', 'english', '英文释义']
    pronun_keywords = ['发音', 'pronunciation', '读音']

    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).lower().strip()
        if col_map['latin_term'] is None and any(k in hl for k in latin_keywords):
            col_map['latin_term'] = i
        elif col_map['chinese_meaning'] is None and any(k in hl for k in chinese_keywords):
            col_map['chinese_meaning'] = i
        elif col_map['english_meaning'] is None and any(k in hl for k in english_keywords):
            col_map['english_meaning'] = i
        elif col_map['pronunciation'] is None and any(k in hl for k in pronun_keywords):
            col_map['pronunciation'] = i

    # 退化：如果没检测到，按顺序猜测
    if col_map['latin_term'] is None:
        col_map['latin_term'] = 0
    if col_map['chinese_meaning'] is None and len(headers) > 1:
        col_map['chinese_meaning'] = 1
    if col_map['english_meaning'] is None and len(headers) > 2:
        col_map['english_meaning'] = 2
    if col_map['pronunciation'] is None and len(headers) > 3:
        col_map['pronunciation'] = 3

    return col_map


def get_cell(row, idx):
    """安全获取单元格值"""
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    return str(val).strip() or None


def import_dictionary(excel_path, dry_run=False):
    """导入词典 Excel 到 SQLite"""
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Excel 文件为空")
        return

    # 检测表头
    headers = rows[0]
    col_map = detect_columns(headers)
    print(f"检测到的列映射:")
    for k, v in col_map.items():
        h = headers[v] if v is not None and v < len(headers) else '?'
        print(f"  {k}: 第 {v} 列 ({h})")

    # 解析数据
    entries = []
    for row in rows[1:]:
        latin = get_cell(row, col_map['latin_term'])
        if not latin:
            continue
        entries.append((
            latin,
            get_cell(row, col_map['chinese_meaning']),
            get_cell(row, col_map['english_meaning']),
            get_cell(row, col_map['pronunciation']),
        ))

    print(f"\n解析到 {len(entries)} 条词典记录")

    if dry_run:
        print(f"\n[dry-run] 前 10 条:")
        for e in entries[:10]:
            print(f"  {e[0]}: {e[1] or '-'} / {e[2] or '-'} [{e[3] or '-'}]")
        return

    # 写入数据库
    conn = sqlite3.connect(str(DB_PATH))
    init_dict_table(conn)

    # 清空已有数据（全量替换）
    conn.execute('DELETE FROM dictionary')

    conn.executemany(
        'INSERT INTO dictionary (latin_term, chinese_meaning, english_meaning, pronunciation) VALUES (?, ?, ?, ?)',
        entries
    )
    conn.commit()
    conn.close()

    print(f"已导入 {len(entries)} 条记录到 {DB_PATH}")

    wb.close()


def main():
    parser = argparse.ArgumentParser(description='植物资料库 - 拉丁词典导入')
    parser.add_argument('excel', help='词典 Excel 文件路径')
    parser.add_argument('--dry-run', action='store_true', help='仅预览，不写入数据库')
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    import_dictionary(excel_path, dry_run=args.dry_run)


if __name__ == '__main__':
    main()
